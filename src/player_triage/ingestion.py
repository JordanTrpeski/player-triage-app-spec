"""CSV and XLSX ingestion with strict header + row validation.

Two entry points:

* :func:`load_csv` — reads the authoritative CSV. This is the canonical format.
* :func:`load_xlsx` — a controlled adapter for the workbook. The adapter reads
  every cell as a string, then delegates the same validation logic used by the
  CSV loader so both formats produce identical :class:`RawMessage` streams.

Both loaders enforce the required-column set, reject duplicated or unexpected
columns, and reject duplicate ``msg_id`` values across rows. Malformed rows are
never silently repaired; they surface as :class:`IngestionError`.

Sanitized errors identify the offending row by ``msg_id`` (when parseable) or
by row number, and never include subject/body/player_id content or any
sensitive value.
"""

from __future__ import annotations

import csv
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Final, Iterable, Iterator

from openpyxl import load_workbook

from .errors import ConfigurationError
from .records import RawMessage

REQUIRED_COLUMNS: Final[tuple[str, ...]] = (
    "msg_id",
    "received_utc",
    "channel",
    "market",
    "player_id",
    "vip_tier",
    "language",
    "subject",
    "body",
)

# Bounds are conservative and align with published service-desk practice. They
# guard against pathological inputs (e.g. multi-MB "message bodies" from a
# corrupt workbook cell) rather than encode business policy about content.
_MAX_SUBJECT_LENGTH: Final[int] = 300
_MAX_BODY_LENGTH: Final[int] = 8_000
_MAX_LANGUAGE_LENGTH: Final[int] = 12
_ALLOWED_CHANNELS: Final[frozenset[str]] = frozenset({"email", "chat"})
_ALLOWED_MARKETS: Final[frozenset[str]] = frozenset(
    {"Ontario", "Malta", "Ireland", "India", "New Zealand"}
)
_MESSAGE_ID_PATTERN: Final[re.Pattern[str]] = re.compile(r"^M\d{2}$")
_PLAYER_ID_PATTERN: Final[re.Pattern[str]] = re.compile(r"^P-\d{5}$")


class IngestionError(ConfigurationError):
    """Raised when the input dataset cannot be safely ingested."""


#: Returns a sanitized error fragment for an unacceptable message identifier,
#: or ``None`` when the identifier is acceptable.
MessageIdValidator = Callable[[str], "str | None"]


def _require_headers(headers: list[str], *, source: Path) -> None:
    seen: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header in seen:
            raise IngestionError(
                component="ingestion",
                message=(
                    f"duplicated header {header!r} at column index {index} "
                    f"(previously at {seen[header]})"
                ),
                path=source,
            )
        seen[header] = index

    missing = tuple(column for column in REQUIRED_COLUMNS if column not in seen)
    if missing:
        raise IngestionError(
            component="ingestion",
            message=f"missing required columns: {list(missing)}",
            path=source,
        )
    unexpected = tuple(column for column in headers if column not in REQUIRED_COLUMNS)
    if unexpected:
        raise IngestionError(
            component="ingestion",
            message=f"unexpected columns present: {list(unexpected)}",
            path=source,
        )


def _parse_timestamp(value: str, *, msg_id: str, source: Path) -> datetime:
    text = value.strip()
    if not text:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: empty received_utc",
            path=source,
        )
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: invalid received_utc timestamp",
            path=source,
        ) from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _benchmark_message_id_error(msg_id: str) -> str | None:
    """Supplied-40 benchmark identifier rule: ``^M\\d{2}$`` (M01-M40).

    This is the accepted benchmark contract and is deliberately unchanged.
    Imported datasets use the wider ``source_message_id`` rule in
    :mod:`player_triage.imported_identifiers` instead.
    """

    if not _MESSAGE_ID_PATTERN.match(msg_id):
        return "msg_id must match ^M\\d{2}$"
    return None


def _validate_and_build(
    row: dict[str, str],
    *,
    source: Path,
    source_row: int,
    source_format: str,
    seen_ids: set[str],
    message_id_error: MessageIdValidator = _benchmark_message_id_error,
) -> RawMessage:
    """Validate one row and build a :class:`RawMessage`, or raise.

    ``message_id_error`` is injected so the imported-data path can apply the
    wider ``^M[0-9]{1,9}$`` rule while reusing byte-for-byte identical
    validation of every other field. The default preserves the supplied-40
    benchmark behaviour exactly.
    """

    msg_id = (row.get("msg_id") or "").strip()
    id_error = message_id_error(msg_id)
    if id_error is not None:
        raise IngestionError(
            component="ingestion",
            message=f"row {source_row}: {id_error}",
            path=source,
        )
    if msg_id in seen_ids:
        raise IngestionError(
            component="ingestion",
            message=f"row {source_row}: duplicate msg_id {msg_id}",
            path=source,
        )

    channel = (row.get("channel") or "").strip()
    if channel not in _ALLOWED_CHANNELS:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: unsupported channel value",
            path=source,
        )
    market = (row.get("market") or "").strip()
    if market not in _ALLOWED_MARKETS:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: unsupported market value",
            path=source,
        )

    language = (row.get("language") or "").strip()
    if not language or len(language) > _MAX_LANGUAGE_LENGTH:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: invalid language length",
            path=source,
        )

    player_id = (row.get("player_id") or "").strip()
    if not _PLAYER_ID_PATTERN.match(player_id):
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: player_id does not match expected internal format",
            path=source,
        )

    subject = row.get("subject") or ""
    body = row.get("body") or ""
    if not subject.strip() and not body.strip():
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: empty subject and body — will fail closed downstream",
            path=source,
        )
    if len(subject) > _MAX_SUBJECT_LENGTH:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: subject exceeds max length ({_MAX_SUBJECT_LENGTH})",
            path=source,
        )
    if len(body) > _MAX_BODY_LENGTH:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: body exceeds max length ({_MAX_BODY_LENGTH})",
            path=source,
        )

    try:
        subject.encode("utf-8")
        body.encode("utf-8")
    except UnicodeEncodeError as exc:
        raise IngestionError(
            component="ingestion",
            message=f"row {msg_id}: non-UTF-8 text detected: {exc.reason}",
            path=source,
        ) from exc

    received_utc = _parse_timestamp(row.get("received_utc") or "", msg_id=msg_id, source=source)

    seen_ids.add(msg_id)
    return RawMessage(
        msg_id=msg_id,
        received_utc=received_utc,
        channel=channel,
        market=market,
        language=language,
        subject=subject,
        body=body,
        player_id=player_id,
        source_format=source_format,
        source_row=source_row,
    )


def _iter_csv_rows(path: Path) -> Iterator[tuple[int, dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        _require_headers(list(headers), source=path)
        for index, row in enumerate(reader, start=2):
            yield index, {key: (value if value is not None else "") for key, value in row.items()}


def load_csv(path: Path | str) -> tuple[RawMessage, ...]:
    """Ingest the authoritative CSV. Returns an immutable tuple of RawMessages."""

    source = Path(path)
    if not source.is_file():
        raise IngestionError(
            component="ingestion", message="CSV input not found", path=source
        )
    seen: set[str] = set()
    messages: list[RawMessage] = []
    for row_number, row in _iter_csv_rows(source):
        messages.append(
            _validate_and_build(
                row,
                source=source,
                source_row=row_number,
                source_format="csv",
                seen_ids=seen,
            )
        )
    return tuple(messages)


def _iter_xlsx_rows(path: Path) -> Iterator[tuple[int, dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise IngestionError(
                component="ingestion",
                message="XLSX workbook has no active worksheet",
                path=path,
            )
        rows_iter: Iterator[tuple[Any, ...]] = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise IngestionError(
                component="ingestion",
                message="XLSX worksheet is empty",
                path=path,
            ) from exc
        headers = [str(value) if value is not None else "" for value in header_row]
        _require_headers(headers, source=path)

        for offset, row in enumerate(rows_iter, start=2):
            if all(cell is None or cell == "" for cell in row):
                # Trailing blank rows are silently skipped, matching CSV behaviour.
                continue
            record = {
                header: ("" if cell is None else str(cell))
                for header, cell in zip(headers, row)
            }
            yield offset, record
    finally:
        workbook.close()


def load_xlsx(path: Path | str) -> tuple[RawMessage, ...]:
    """Ingest the XLSX workbook through a controlled adapter."""

    source = Path(path)
    if not source.is_file():
        raise IngestionError(
            component="ingestion", message="XLSX input not found", path=source
        )
    seen: set[str] = set()
    messages: list[RawMessage] = []
    for row_number, row in _iter_xlsx_rows(source):
        messages.append(
            _validate_and_build(
                row,
                source=source,
                source_row=row_number,
                source_format="xlsx",
                seen_ids=seen,
            )
        )
    return tuple(messages)


def load(path: Path | str) -> tuple[RawMessage, ...]:
    """Dispatch to the appropriate loader based on file suffix."""

    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return load_csv(source)
    if suffix == ".xlsx":
        return load_xlsx(source)
    raise IngestionError(
        component="ingestion",
        message=f"unsupported input format {suffix!r}",
        path=source,
    )


def message_ids(messages: Iterable[RawMessage]) -> tuple[str, ...]:
    return tuple(m.msg_id for m in messages)
