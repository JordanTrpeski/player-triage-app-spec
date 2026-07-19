"""Sanitized read/review services for the local control console."""

from __future__ import annotations

import csv
import io
import json
import re
import tempfile
from dataclasses import asdict
from importlib.metadata import PackageNotFoundError, version as package_version
from pathlib import Path
from typing import Any, Final, Mapping, Sequence

from .configuration_manager import ConfigurationManager
from .console_contracts import (
    AuditView,
    DashboardSnapshot,
    ImportedRunDetail,
    ImportedRunSummary,
    ImportPreview,
    ImportRunView,
    MessageView,
    VersionView,
)
from .errors import ConfigurationError
from .operational import (
    AUDIT_FILENAME,
    DECISIONS_JSONL_FILENAME,
    MANIFEST_FILENAME,
    append_human_override_decision,
    verify_run_artifacts,
)
from .routing import load_routing_map


_ROUTING = load_routing_map()


#: Shape of an internally generated imported-run identifier. Used to reject
#: crafted values before they are joined to a filesystem path.
_IMPORT_RUN_ID_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^irun-[0-9A-Za-z]+-[0-9a-f]{12}$"
)


#: Longest subject fragment shown in the import preview.
_PREVIEW_SUBJECT_CHARS: Final[int] = 48


#: Imported-run statuses whose artifacts are complete enough to be selected and
#: displayed. ``started`` means the run was interrupted and ``failed`` means it
#: never produced decisions, so neither is offered as a dashboard dataset.
SELECTABLE_IMPORT_STATUSES: Final[tuple[str, ...]] = (
    "completed",
    "completed_with_errors",
)


#: Columns surfaced in the imported decision table, in display order. Every one
#: is a structured decision field or an internally generated identifier; the
#: source subject, body and player identifier are absent from ``decisions.csv``
#: by construction and are never reintroduced here.
_IMPORTED_DECISION_COLUMNS: Final[tuple[str, ...]] = (
    "source_message_id",
    "case_ref",
    "category",
    "intent",
    "priority",
    "route",
    "assigned_team",
    "human_review_required",
    "processing_status",
    "model_called",
)


#: Fields the dashboard summarizes as distributions for an imported run.
_IMPORTED_DISTRIBUTION_FIELDS: Final[tuple[str, ...]] = (
    "category",
    "priority",
    "route",
    "assigned_team",
)


def _safe_int(value: object) -> int:
    """Coerce a manifest value to ``int``, treating anything unusable as zero.

    A hand-edited or partially written manifest can carry a string, ``null`` or
    a nested object where a count belongs. Counts are presentational here, so a
    bad value degrades that one number rather than failing the whole page.
    """

    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, (str, float)):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0
    return 0


def _load_manifest(path: Path) -> Mapping[str, Any] | None:
    """Read a run manifest, returning ``None`` when it is unusable.

    Covers the absent, unreadable, non-JSON and wrong-shape cases in one place
    so every caller degrades identically instead of raising.
    """

    if not path.is_file():
        return None
    try:
        with path.open(encoding="utf-8") as handle:
            loaded = json.load(handle)
    except (OSError, ValueError, UnicodeDecodeError):
        return None
    if not isinstance(loaded, Mapping):
        return None
    return loaded


def _counts_from_rows(rows: Sequence[Mapping[str, str]], field: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        key = str(row.get(field, "") or "—")
        counts[key] = counts.get(key, 0) + 1
    return dict(sorted(counts.items()))


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open(newline="", encoding="utf-8") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _truncate(value: str) -> str:
    text = " ".join(value.split())
    if len(text) <= _PREVIEW_SUBJECT_CHARS:
        return text
    return text[:_PREVIEW_SUBJECT_CHARS] + "…"


def _read_structure(payload: bytes, suffix: str) -> tuple[list[str], list[dict[str, str]]]:
    """Return (headers, rows) for preview. Structure only, never validated."""

    if suffix == ".csv":
        text = payload.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = [
            {key: (value if value is not None else "") for key, value in row.items()}
            for row in reader
        ]
        return headers, rows

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return [], []
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            return [], []
        headers = [str(v) if v is not None else "" for v in header_row]
        rows = []
        for raw in iterator:
            if all(cell is None or cell == "" for cell in raw):
                continue
            rows.append(
                {
                    header: ("" if cell is None else str(cell))
                    for header, cell in zip(headers, raw)
                }
            )
        return headers, rows
    finally:
        workbook.close()


class ConsoleServiceError(ConfigurationError):
    """Sanitized console read/review failure."""


class ConsoleService:
    """Single typed facade consumed by Streamlit pages."""

    def __init__(
        self,
        app_root: Path,
        *,
        state_root: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.app_root = app_root.resolve()
        self.output_root = (
            output_root.resolve()
            if output_root is not None
            else (self.app_root / "output").resolve()
        )
        self.configuration = ConfigurationManager(self.app_root, state_root)

    def dashboard(self) -> DashboardSnapshot:
        run_dir = self.latest_run_dir()
        if run_dir is None:
            settings = self.configuration.settings()
            return DashboardSnapshot(
                active_policy_version=str(settings["active_policy_version"]),
                application_version=_application_version(),
                runtime_mode="rules_only",
                model_status="rejected / disabled",
                model_calls=0,
                kill_switch_enabled=bool(settings["model_kill_switch_enabled"]),
                latest_run_id="none",
                latest_run_status="no verified run",
                canonical_digest="unavailable",
                counts={"input": 0, "success": 0, "failure": 0, "bypass": 0},
                distributions={},
                manual_review_rate=0.0,
                specialist_rate=0.0,
                official_gates_passed=0,
                official_gate_count=15,
                locked_gates_passed=0,
                locked_gate_count=26,
                core_mismatch_count=0,
                diagnostic_difference_count=0,
                p50_latency_ms=0.0,
                p95_latency_ms=0.0,
                messages_per_second=0.0,
                replay_900_seconds=0.0,
            )
        config = self.configuration.load_active_config()
        verify_run_artifacts(config, run_dir)
        manifest = _read_json(run_dir / MANIFEST_FILENAME)
        decisions = _read_jsonl(run_dir / DECISIONS_JSONL_FILENAME)
        evaluation = self._evaluation_documents()
        supplied = _supplied_result(evaluation.get("dataset_results", {}))
        safety = evaluation.get("safety", {})
        performance = evaluation.get("performance", {})
        capacity = evaluation.get("capacity", {})
        route_counts = _counts(decisions, "route")
        specialist = _ROUTING.constants.specialist
        manual_count = sum(
            decision.get("human_review_required") is True for decision in decisions
        )
        settings = self.configuration.settings()
        return DashboardSnapshot(
            active_policy_version=str(settings["active_policy_version"]),
            application_version=_application_version(),
            runtime_mode="rules_only",
            model_status="rejected / disabled",
            model_calls=sum(decision.get("model_called") is True for decision in decisions),
            kill_switch_enabled=bool(settings["model_kill_switch_enabled"]),
            latest_run_id=str(manifest.get("run_id", "unavailable")),
            latest_run_status=str(manifest.get("status", "unavailable")),
            canonical_digest=str(manifest.get("canonical_decision_sha256", "unavailable")),
            counts={
                "input": int(manifest.get("message_count", 0)),
                "success": int(manifest.get("success_count", 0)),
                "failure": int(manifest.get("failure_count", 0)),
                "bypass": int(manifest.get("bypass_count", 0)),
            },
            distributions={
                "category": _counts(decisions, "category"),
                "priority": _counts(decisions, "priority"),
                "route": route_counts,
                "assigned_team": _counts(decisions, "assigned_team"),
            },
            manual_review_rate=manual_count / len(decisions) if decisions else 0.0,
            specialist_rate=route_counts.get(specialist, 0) / len(decisions)
            if decisions
            else 0.0,
            official_gates_passed=sum(
                bool(item.get("passed"))
                for item in safety.get("results", [])
                if str(item.get("gate_id", "")).startswith("S")
            ),
            official_gate_count=15,
            locked_gates_passed=sum(
                bool(item.get("passed")) for item in safety.get("results", [])
            ),
            locked_gate_count=26,
            core_mismatch_count=len(supplied.get("mismatches", [])),
            diagnostic_difference_count=len(
                supplied.get("diagnostic_differences", [])
            ),
            p50_latency_ms=float(performance.get("per_message_median_latency_ms", 0)),
            p95_latency_ms=float(performance.get("per_message_p95_latency_ms", 0)),
            messages_per_second=float(performance.get("messages_per_second", 0)),
            replay_900_seconds=float(
                capacity.get("full_day_replay_seconds_at_measured_throughput", 0)
            ),
        )

    def messages(
        self, filters: Mapping[str, object] | None = None
    ) -> Sequence[MessageView]:
        run_dir = self.latest_run_dir()
        if run_dir is None:
            return ()
        decisions = _read_jsonl(run_dir / DECISIONS_JSONL_FILENAME)
        events = {
            str(item.get("message_id")): item
            for item in _read_jsonl(run_dir / AUDIT_FILENAME)
            if item.get("event_type") == "decision"
        }
        mismatch_records = self._mismatch_records()
        by_message: dict[str, list[dict[str, Any]]] = {}
        for item in mismatch_records:
            if item.get("dataset_name") == "supplied-40":
                by_message.setdefault(str(item.get("message_id")), []).append(item)
        manifest = _read_json(run_dir / MANIFEST_FILENAME)
        views: list[MessageView] = []
        for decision in decisions:
            message_id = str(decision.get("message_id"))
            differences = tuple(by_message.get(message_id, ()))
            event = events.get(message_id, {})
            view = MessageView(
                message_id=message_id,
                decision=decision,
                expected_actual=differences,
                core_mismatch=any(
                    item.get("adjudication_status") != "diagnostic_non_baseline_difference"
                    for item in differences
                ),
                diagnostic_difference=any(
                    item.get("adjudication_status") == "diagnostic_non_baseline_difference"
                    for item in differences
                ),
                rules_triggered=tuple(
                    str(value) for value in event.get("payload", {}).get("rules_triggered", ())
                ),
                decision_path=str(event.get("payload", {}).get("decision_path", "")),
                audit_event_id=str(event.get("event_id", "")),
                configuration_version=str(manifest.get("policy_bundle_version", "")),
            )
            if _matches_message_filters(view, filters or {}):
                views.append(view)
        return tuple(views)

    def review_queue(self) -> Sequence[MessageView]:
        return tuple(
            view
            for view in self.messages()
            if view.decision.get("human_review_required") is True
            or bool(view.decision.get("missing_context"))
            or view.core_mismatch
            or view.diagnostic_difference
        )

    def submit_override(
        self,
        message_id: str,
        proposed: Mapping[str, Any],
        reason_code: str,
        actor_label: str,
    ) -> str:
        run_dir = self.latest_run_dir()
        if run_dir is None:
            raise self._error("human_override", "no verified run is available")
        original = next(
            (view.decision for view in self.messages() if view.message_id == message_id),
            None,
        )
        if original is None:
            raise self._error("human_override", "message was not found")
        after = dict(original)
        allowed = {
            "category",
            "intent",
            "secondary_intents",
            "priority",
            "route",
            "assigned_team",
            "secondary_teams",
            "risk_flags",
            "reason_codes",
        }
        unknown = set(proposed) - allowed
        if unknown:
            raise self._error("human_override", "override contains an unsupported field")
        after.update(proposed)
        after["decision_basis"] = "human_override"
        return append_human_override_decision(
            self.configuration.load_active_config(),
            run_dir=run_dir,
            message_id=message_id,
            reason_code=reason_code,
            after=after,
            actor_label=actor_label,
        )

    def audit_events(
        self, filters: Mapping[str, object] | None = None
    ) -> Sequence[AuditView]:
        documents: list[dict[str, Any]] = []
        run_dir = self.latest_run_dir()
        if run_dir is not None:
            documents.extend(_read_jsonl(run_dir / AUDIT_FILENAME))
        documents.extend(self.configuration.control_audit_events())
        output: list[AuditView] = []
        for event in documents:
            view = AuditView(
                event_id=str(event.get("event_id", "")),
                run_id=str(event.get("run_id", "")),
                message_id=(
                    str(event["message_id"])
                    if event.get("message_id") is not None
                    else None
                ),
                event_type=str(event.get("event_type", "")),
                occurred_at=str(event.get("occurred_at", "")),
                configuration_version=str(event.get("configuration_version", "")),
                actor=dict(event.get("actor", {})),
                payload=dict(event.get("payload", {})),
            )
            if _matches_audit_filters(view, filters or {}):
                output.append(view)
        return tuple(sorted(output, key=lambda item: item.occurred_at, reverse=True))

    def versions(self) -> Sequence[VersionView]:
        active = self.configuration.active_state()["version_id"]
        return tuple(
            VersionView(
                version_id=str(item.get("version_id", "")),
                parent_version_id=(
                    str(item["parent_version_id"])
                    if item.get("parent_version_id") is not None
                    else None
                ),
                status=str(item.get("status", "")),
                actor=str(item.get("actor", "")),
                change_reason=str(item.get("change_reason", "")),
                bundle_digest=str(item.get("bundle_digest", "")),
                validation_passed=_optional_bool(item.get("validation_passed")),
                regression_passed=_optional_bool(item.get("regression_passed")),
                gates_passed=_optional_bool(item.get("gates_passed")),
                activated_at=(
                    str(item["activated_at"])
                    if item.get("activated_at") is not None
                    else None
                ),
                rollback_available=str(item.get("version_id")) != str(active)
                and str(item.get("status")) != "draft",
                summary=str(item.get("summary", "")),
            )
            for item in self.configuration.versions()
        )

    def settings(self) -> Mapping[str, Any]:
        return self.configuration.settings()

    def policy_components(self) -> Mapping[str, Any]:
        config = self.configuration.load_active_config()
        editability = config.component("ui_editability").get("components", {})
        output: dict[str, Any] = {}
        for component in (
            "policy_rules",
            "baseline_intent_rules",
            "derived_refinement_rules",
            "redaction_policy",
            "market_overlays",
            "auto_response_templates",
            "rationale_templates",
            "semantic_constraints",
            "model_configuration",
        ):
            document = dict(config.component(component))
            if component == "model_configuration":
                for key in ("local_path_reference", "approved_model_id", "sha256"):
                    document.pop(key, None)
            output[component] = {
                "version": document.get("version"),
                "digest": config.component_digest(component),
                "ui": editability.get(component, {"normal_ui": "read_only"}),
                "document": document,
            }
        return output

    def evaluation_documents(self) -> Mapping[str, Any]:
        return self._evaluation_documents()

    def safe_downloads(self) -> Mapping[str, bytes]:
        allowed = (
            "evaluation_summary.json",
            "mismatch_report.csv",
            "confusion_matrix.csv",
            "safety_gate_results.json",
            "performance_results.json",
            "capacity_estimate.json",
            "human_review_workload.json",
        )
        output: dict[str, bytes] = {}
        for name in allowed:
            path = self.output_root / name
            if path.is_file():
                output[name] = path.read_bytes()
        return output

    def latest_run_dir(self) -> Path | None:
        candidates: list[tuple[str, Path]] = []
        if not self.output_root.is_dir():
            return None
        for manifest_path in self.output_root.rglob(MANIFEST_FILENAME):
            directory = manifest_path.parent
            if not (directory / DECISIONS_JSONL_FILENAME).is_file():
                continue
            try:
                manifest = _read_json(manifest_path)
            except ConsoleServiceError:
                continue
            candidates.append((str(manifest.get("completed_at", "")), directory))
        return max(candidates, key=lambda item: item[0])[1] if candidates else None

    def _evaluation_documents(self) -> dict[str, Any]:
        names = {
            "summary": "evaluation_summary.json",
            "dataset_results": "dataset_results.json",
            "safety": "safety_gate_results.json",
            "performance": "performance_results.json",
            "capacity": "capacity_estimate.json",
            "workload": "human_review_workload.json",
            "cost": "cost_estimate.json",
            "audit_reconstruction": "audit_reconstruction.json",
        }
        output: dict[str, Any] = {}
        for key, name in names.items():
            path = self.output_root / name
            if path.is_file():
                try:
                    output[key] = _read_json(path)
                except ConsoleServiceError:
                    continue
        return output

    def _mismatch_records(self) -> list[dict[str, Any]]:
        path = self.output_root / "mismatch_report.jsonl"
        return _read_jsonl(path) if path.is_file() else []

    # -- imported runs -----------------------------------------------------

    def run_import(
        self,
        payload: bytes,
        *,
        display_name: str,
        collision_mode: str = "error",
    ) -> ImportRunView:
        """Process an uploaded batch into an isolated, application-owned run.

        The uploaded bytes are staged in a temporary directory that is removed
        afterwards. The operator never selects a server-side destination: runs
        are always written beneath ``output/imported_runs``.
        """

        from .config import load_app_config
        from .imported_runs import (
            IMPORTED_RUNS_DIRNAME,
            VALIDATION_ERRORS_FILENAME,
            run_imported_batch,
            sanitize_display_name,
        )

        safe_name = sanitize_display_name(display_name)
        suffix = Path(safe_name).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise self._error("import", "unsupported input format")

        config = load_app_config(self.app_root)
        with tempfile.TemporaryDirectory(prefix="player-triage-import-") as staging:
            staged = Path(staging) / f"upload{suffix}"
            staged.write_bytes(payload)
            result = run_imported_batch(
                config,
                staged,
                display_name=safe_name,
                collision_mode=collision_mode,
                # Always the service's own application-owned root, so the
                # console and its artifact reads agree and the operator never
                # influences the destination.
                output_root=self.output_root / IMPORTED_RUNS_DIRNAME,
            )

        rejected = _read_csv_rows(result.run_dir / VALIDATION_ERRORS_FILENAME)
        return ImportRunView(
            run_id=result.run_id,
            status=result.status,
            policy_version=result.policy_version,
            rows_seen=result.rows_seen,
            rows_accepted=result.rows_accepted,
            rows_rejected=result.rows_rejected,
            rows_processed=result.rows_processed,
            rows_failed=result.rows_failed,
            decision_digest=result.decision_digest,
            model_calls=result.model_calls,
            rejected_rows=tuple(rejected),
        )

    def read_import_artifact(self, run_id: str, filename: str) -> bytes | None:
        """Return one artifact from an imported run, or ``None`` if absent.

        ``run_id`` and ``filename`` are both checked against allow-lists so a
        crafted value cannot read outside the imported-run root.
        """

        from .imported_runs import (
            AUDIT_JSONL_FILENAME,
            DECISIONS_CSV_FILENAME,
            IMPORTED_RUNS_DIRNAME,
            PROCESSING_SUMMARY_FILENAME,
            RUN_MANIFEST_FILENAME,
            VALIDATION_ERRORS_FILENAME,
        )

        allowed = {
            DECISIONS_CSV_FILENAME,
            AUDIT_JSONL_FILENAME,
            VALIDATION_ERRORS_FILENAME,
            RUN_MANIFEST_FILENAME,
            PROCESSING_SUMMARY_FILENAME,
        }
        if filename not in allowed:
            return None
        if not _IMPORT_RUN_ID_PATTERN.match(run_id):
            return None

        root = (self.output_root / IMPORTED_RUNS_DIRNAME).resolve()
        target = (root / run_id / filename).resolve()
        if not target.is_relative_to(root) or not target.is_file():
            return None
        return target.read_bytes()

    def import_template_csv(self) -> bytes:
        """A downloadable CSV template for the fixed import contract.

        Header row plus one clearly synthetic example row. Contains no real
        player data and no sensitive values.
        """

        from .ingestion import REQUIRED_COLUMNS

        example = {
            "msg_id": "M1",
            "received_utc": "2026-01-01T09:00:00Z",
            "channel": "email",
            "market": "Ontario",
            "player_id": "P-00000",
            "vip_tier": "none",
            "language": "en",
            "subject": "EXAMPLE ROW - replace or delete before import",
            "body": "Synthetic example. No real player data.",
        }
        buffer = io.StringIO()
        writer = csv.DictWriter(
            buffer, fieldnames=list(REQUIRED_COLUMNS), lineterminator="\n"
        )
        writer.writeheader()
        writer.writerow(example)
        return buffer.getvalue().encode("utf-8")

    def preview_import(
        self, payload: bytes, *, display_name: str, max_rows: int = 5
    ) -> ImportPreview:
        """Inspect an uploaded batch before processing it.

        Reads structure only. Nothing is classified, no run directory is
        created and no artifact is written. Subjects are truncated and bodies
        are never included.
        """

        from .imported_runs import sanitize_display_name
        from .ingestion import REQUIRED_COLUMNS

        safe_name = sanitize_display_name(display_name)
        suffix = Path(safe_name).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise self._error("import", "unsupported input format")

        headers, rows = _read_structure(payload, suffix)
        required = set(REQUIRED_COLUMNS)
        present = set(headers)

        sample: list[dict[str, str]] = []
        for row in rows[:max_rows]:
            subject = row.get("subject", "")
            sample.append(
                {
                    "msg_id": row.get("msg_id", ""),
                    "received_utc": row.get("received_utc", ""),
                    "channel": row.get("channel", ""),
                    "market": row.get("market", ""),
                    "language": row.get("language", ""),
                    "subject_preview": _truncate(subject),
                }
            )

        return ImportPreview(
            display_name=safe_name,
            detected_format=suffix.lstrip("."),
            row_count=len(rows),
            detected_columns=tuple(headers),
            missing_columns=tuple(c for c in REQUIRED_COLUMNS if c not in present),
            unexpected_columns=tuple(c for c in headers if c not in required),
            sample_rows=tuple(sample),
            truncated=len(rows) > max_rows,
        )

    def recent_imported_runs(self, limit: int = 20) -> Sequence[ImportedRunSummary]:
        """Safe metadata for recent imported runs, newest first.

        Reads only each run's manifest. No message content is exposed.
        """

        from .imported_runs import IMPORTED_RUNS_DIRNAME, RUN_MANIFEST_FILENAME

        root = self.output_root / IMPORTED_RUNS_DIRNAME
        if not root.is_dir():
            return ()

        summaries: list[ImportedRunSummary] = []
        for entry in sorted(root.iterdir(), reverse=True):
            if not entry.is_dir() or not _IMPORT_RUN_ID_PATTERN.match(entry.name):
                continue
            manifest_path = entry / RUN_MANIFEST_FILENAME
            if not manifest_path.is_file():
                continue
            doc = _load_manifest(manifest_path)
            if doc is None:
                # A partially written or corrupt manifest must not break the
                # list; the run simply does not appear.
                continue
            summaries.append(
                ImportedRunSummary(
                    run_id=str(doc.get("run_id", entry.name)),
                    status=str(doc.get("status", "unknown")),
                    started_at=str(doc.get("started_at", "")),
                    completed_at=(
                        str(doc["completed_at"])
                        if doc.get("completed_at") is not None
                        else None
                    ),
                    source_filename_sanitized=str(
                        doc.get("source_filename_sanitized", "")
                    ),
                    rows_seen=_safe_int(doc.get("rows_seen")),
                    rows_processed=_safe_int(doc.get("rows_processed")),
                    rows_rejected=_safe_int(doc.get("rows_rejected")),
                    policy_version=str(doc.get("policy_version", "")),
                    decision_digest=(
                        str(doc["decision_digest"])
                        if doc.get("decision_digest")
                        else None
                    ),
                )
            )
            if len(summaries) >= limit:
                break
        return tuple(summaries)

    def selectable_imported_runs(
        self, limit: int = 20
    ) -> Sequence[ImportedRunSummary]:
        """Imported runs complete enough to open on the dashboard, newest first.

        Interrupted (``started``) and ``failed`` runs are excluded: neither has
        a trustworthy decision set to display.
        """

        return tuple(
            run
            for run in self.recent_imported_runs(limit=limit)
            if run.status in SELECTABLE_IMPORT_STATUSES
        )

    def imported_run_detail(self, run_id: str) -> ImportedRunDetail | None:
        """Full dashboard view of one imported run, or ``None`` if unusable.

        Returns ``None`` — rather than raising — for an unknown, crafted,
        unfinished or corrupt run, so a damaged directory degrades to "not
        selectable" instead of breaking the dashboard.
        """

        from .imported_runs import (
            DECISIONS_CSV_FILENAME,
            IMPORTED_RUNS_DIRNAME,
            RUN_MANIFEST_FILENAME,
        )

        if not _IMPORT_RUN_ID_PATTERN.match(run_id):
            return None

        root = (self.output_root / IMPORTED_RUNS_DIRNAME).resolve()
        run_dir = (root / run_id).resolve()
        if not run_dir.is_relative_to(root) or not run_dir.is_dir():
            return None

        doc = _load_manifest(run_dir / RUN_MANIFEST_FILENAME)
        if doc is None:
            return None
        if str(doc.get("status", "")) not in SELECTABLE_IMPORT_STATUSES:
            return None

        try:
            rows = _read_csv_rows(run_dir / DECISIONS_CSV_FILENAME)
        except (OSError, ValueError, UnicodeDecodeError):
            rows = []

        return ImportedRunDetail(
            run_id=str(doc.get("run_id", run_id)),
            status=str(doc.get("status", "unknown")),
            source_filename_sanitized=str(doc.get("source_filename_sanitized", "")),
            started_at=str(doc.get("started_at", "")),
            completed_at=(
                str(doc["completed_at"])
                if doc.get("completed_at") is not None
                else None
            ),
            rows_seen=_safe_int(doc.get("rows_seen")),
            rows_accepted=_safe_int(doc.get("rows_accepted")),
            rows_rejected=_safe_int(doc.get("rows_rejected")),
            rows_processed=_safe_int(doc.get("rows_processed")),
            rows_failed=_safe_int(doc.get("rows_failed")),
            policy_version=str(doc.get("policy_version", "")),
            model_calls=_safe_int(doc.get("model_calls")),
            decision_digest=(
                str(doc["decision_digest"]) if doc.get("decision_digest") else None
            ),
            distributions={
                field: _counts_from_rows(rows, field)
                for field in _IMPORTED_DISTRIBUTION_FIELDS
            },
            decisions=tuple(
                {
                    column: str(row.get(column, ""))
                    for column in _IMPORTED_DECISION_COLUMNS
                }
                for row in rows
            ),
        )

    def walkthrough_overview(self) -> Mapping[str, Any]:
        """Static orientation facts for the walkthrough page."""

        from .evaluation_service import ACCEPTED_CANONICAL_DIGEST

        settings = self.configuration.settings()
        return {
            "policy_version": settings.get("configuration_version", "policy-3.3.1"),
            "canonical_digest": ACCEPTED_CANONICAL_DIGEST,
            "category_agreement": "40/40",
            "intent_agreement": "39/40",
            "processing_mode": "rules_only",
            "model_calls": 0,
        }

    @staticmethod
    def _error(component: str, message: str) -> ConsoleServiceError:
        return ConsoleServiceError(component=component, message=message)


def _matches_message_filters(view: MessageView, filters: Mapping[str, object]) -> bool:
    decision = view.decision
    scalar_fields = {
        "message_id": view.message_id,
        "market": decision.get("market"),
        "language": decision.get("language"),
        "category": decision.get("category"),
        "intent": decision.get("intent"),
        "priority": decision.get("priority"),
        "route": decision.get("route"),
        "assigned_team": decision.get("assigned_team"),
        "model_eligibility": decision.get("model_eligibility"),
        "model_bypass_reason": decision.get("model_bypass_reason"),
        "human_review_required": decision.get("human_review_required"),
        "configuration_version": view.configuration_version,
        "core_mismatch": view.core_mismatch,
        "diagnostic_difference": view.diagnostic_difference,
    }
    list_fields = {
        "secondary_team": decision.get("secondary_teams", ()),
        "risk_flag": decision.get("risk_flags", ()),
        "reason_code": decision.get("reason_codes", ()),
    }
    for key, expected in filters.items():
        if expected in (None, "", (), []):
            continue
        if key in scalar_fields and scalar_fields[key] != expected:
            return False
        if key in list_fields and expected not in list_fields[key]:
            return False
    return True


def _matches_audit_filters(view: AuditView, filters: Mapping[str, object]) -> bool:
    scalar = {
        "event_id": view.event_id,
        "run_id": view.run_id,
        "message_id": view.message_id,
        "event_type": view.event_type,
        "configuration_version": view.configuration_version,
    }
    serialized = json.dumps(asdict(view), sort_keys=True)
    for key, expected in filters.items():
        if expected in (None, ""):
            continue
        if key in scalar and scalar[key] != expected:
            return False
        if key in {"rule_id", "reason_code", "actor", "component"} and str(expected) not in serialized:
            return False
        if key == "date_from" and view.occurred_at < str(expected):
            return False
        if key == "date_to" and view.occurred_at > str(expected):
            return False
    return True


def _read_json(path: Path) -> dict[str, Any]:
    try:
        document = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ConsoleServiceError(component="console_read", message="structured artifact is unavailable") from exc
    if not isinstance(document, dict):
        raise ConsoleServiceError(component="console_read", message="structured artifact is invalid")
    return document


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError as exc:
        raise ConsoleServiceError(component="console_read", message="structured artifact is unavailable") from exc
    for line in lines:
        if not line.strip():
            continue
        try:
            item = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ConsoleServiceError(component="console_read", message="structured artifact is invalid") from exc
        if not isinstance(item, dict):
            raise ConsoleServiceError(component="console_read", message="structured artifact is invalid")
        output.append(item)
    return output


def _counts(decisions: Sequence[Mapping[str, Any]], field: str) -> dict[str, int]:
    output: dict[str, int] = {}
    for decision in decisions:
        key = str(decision.get(field))
        output[key] = output.get(key, 0) + 1
    return dict(sorted(output.items()))


def _supplied_result(document: Mapping[str, Any]) -> Mapping[str, Any]:
    return next(
        (
            item
            for item in document.get("results", ())
            if isinstance(item, Mapping) and item.get("dataset_name") == "supplied-40"
        ),
        {},
    )


def _optional_bool(value: object) -> bool | None:
    return value if isinstance(value, bool) else None


def _application_version() -> str:
    try:
        return package_version("player_triage")
    except PackageNotFoundError:
        return "uninstalled"
